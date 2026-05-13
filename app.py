import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =========================
# ページ設定
# =========================
st.set_page_config(
    page_title="健康チェックWebアプリ",
    page_icon="🩺",
    layout="wide",
)


# =========================
# ログイン設定
# =========================
USERS = {
    "kanri": {
        "password": "rui",
        "role": "admin",
        "label": "管理者",
    },
    "staff": {
        "password": "rui",
        "role": "staff",
        "label": "職員",
    },
}


def login_check():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    if "role" not in st.session_state:
        st.session_state.role = None

    if "user_label" not in st.session_state:
        st.session_state.user_label = ""

    if st.session_state.logged_in:
        return True

    st.markdown(
        """
        <div style='text-align:center; padding:24px;'>
            <h1 style='color:#2E7D32;'>健康チェック管理システム</h1>
            <p style='color:#666;'>利用者様の健康記録を安全に管理します</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown("### ログイン")
        input_id = st.text_input("ID")
        input_password = st.text_input("パスワード", type="password")

        if st.button("ログイン", use_container_width=True):
            user = USERS.get(input_id)
            if user and input_password == user["password"]:
                st.session_state.logged_in = True
                st.session_state.role = user["role"]
                st.session_state.user_label = user["label"]
                st.rerun()
            else:
                st.error("IDまたはパスワードが違います。")


    return False


def logout_button():
    with st.sidebar:
        st.caption(f"ログイン中：{st.session_state.user_label}")
        if st.button("ログアウト"):
            st.session_state.logged_in = False
            st.session_state.role = None
            st.session_state.user_label = ""
            st.rerun()


if not login_check():
    st.stop()


# =========================
# デザイン
# =========================
def apply_design():
    if st.session_state.role == "staff":
        bg = "#FFFDF7"
        accent = "#D97A6A"
        box = "#FFF1E8"
    else:
        bg = "#F4F6F9"
        accent = "#1F4E79"
        box = "#EAF1F8"

    st.markdown(
        f"""
        <style>
        .stApp {{
            background-color: {bg};
        }}
        h1, h2, h3 {{
            color: {accent};
        }}
        .info-box {{
            background: {box};
            padding: 14px 18px;
            border-radius: 14px;
            border: 1px solid rgba(0,0,0,0.08);
            margin-bottom: 12px;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


apply_design()


# =========================
# 保存先設定
# =========================
APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
REPORT_DIR = APP_DIR / "reports"
DATA_FILE = DATA_DIR / "健康チェック入力データ.xlsx"
USER_FILE = DATA_DIR / "利用者マスタ.xlsx"
SHEET_NAME = "入力データ"
USER_SHEET = "利用者マスタ"


DEFAULT_USERS = [
    "さくら様",
    "谷様",
    "磯崎様",
    "川上様",
    "和波様",
    "桜井様",
    "國枝様",
    "中野様",
    "山口様",
]


ASSESSMENT_COLUMNS = [
    "記録日",
    "利用者名",
    "体温",
    "血圧上",
    "血圧下",
    "脈拍",
    "SpO2",
    "体重",
    "朝食摂取率",
    "昼食摂取率",
    "夕食摂取率",
    "排尿回数",
    "排便回数",
    "午前尿量",
    "午前尿性状",
    "午前便量",
    "午前便性状",
    "午後尿量",
    "午後尿性状",
    "午後便量",
    "午後便性状",
    "夕方尿量",
    "夕方尿性状",
    "夕方便量",
    "夕方便性状",
    "夜尿量",
    "夜尿性状",
    "夜便量",
    "夜便性状",
    "深夜尿量",
    "深夜尿性状",
    "深夜便量",
    "深夜便性状",
    "朝方尿量",
    "朝方尿性状",
    "朝方便量",
    "朝方便性状",
    "家族共有メモ",
    "気になる変化",
    "登録日時",
    "入力者",
]


EXCRETION_SLOTS = [
    ("午前", "9時〜12時"),
    ("午後", "12時〜15時"),
    ("夕方", "15時〜17時"),
    ("夜", "18時〜22時"),
    ("深夜", "22時〜5時"),
    ("朝方", "5時〜8時"),
]

URINE_AMOUNT_OPTIONS = ["なし", "少", "中", "大"]
URINE_TYPE_OPTIONS = ["普通尿", "濃縮尿"]
STOOL_AMOUNT_OPTIONS = ["なし", "少", "中", "大"]
STOOL_TYPE_OPTIONS = ["普通便", "下痢便", "水様便"]


COLUMNS = [
    "記録日",
    "利用者名",
    "体温",
    "血圧上",
    "血圧下",
    "脈拍",
    "SpO2",
    "体重",
    "朝食摂取率",
    "昼食摂取率",
    "夕食摂取率",
    "排尿回数",
    "排便回数",
    "家族共有メモ",
    "気になる変化",
    "登録日時",
    "入力者",
]


# =========================
# 基本関数
# =========================
def ensure_dirs():
    DATA_DIR.mkdir(exist_ok=True)
    REPORT_DIR.mkdir(exist_ok=True)


def ensure_user_file():
    ensure_dirs()
    if not USER_FILE.exists():
        data = {
            "利用者名": DEFAULT_USERS,
            "表示": ["表示"] * len(DEFAULT_USERS),
        }
        for col in ASSESSMENT_COLUMNS:
            data[col] = [""] * len(DEFAULT_USERS)

        df = pd.DataFrame(data)
        df.to_excel(USER_FILE, index=False, sheet_name=USER_SHEET)


def load_users(include_hidden=False):
    ensure_user_file()

    try:
        df = pd.read_excel(USER_FILE, sheet_name=USER_SHEET)
    except Exception:
        data = {
            "利用者名": DEFAULT_USERS,
            "表示": ["表示"] * len(DEFAULT_USERS),
        }
        for col in ASSESSMENT_COLUMNS:
            data[col] = [""] * len(DEFAULT_USERS)
        df = pd.DataFrame(data)

    if "利用者名" not in df.columns:
        df["利用者名"] = DEFAULT_USERS

    if "表示" not in df.columns:
        df["表示"] = "表示"

    for col in ASSESSMENT_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df["利用者名"] = df["利用者名"].astype(str).str.strip()
    df = df[df["利用者名"] != ""].drop_duplicates(
        subset=["利用者名"],
        keep="first",
    )

    if not include_hidden:
        df = df[df["表示"].fillna("表示") == "表示"]

    return df.reset_index(drop=True)


def save_users(df):
    ensure_dirs()

    df = df.copy()

    if "利用者名" not in df.columns:
        df["利用者名"] = ""

    if "表示" not in df.columns:
        df["表示"] = "表示"

    for col in ASSESSMENT_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df["利用者名"] = df["利用者名"].astype(str).str.strip()
    df = df[df["利用者名"] != ""].drop_duplicates(
        subset=["利用者名"],
        keep="first",
    )

    ordered_cols = ["利用者名", "表示"] + ASSESSMENT_COLUMNS
    for col in ordered_cols:
        if col not in df.columns:
            df[col] = ""

    df = df[ordered_cols]
    df.to_excel(USER_FILE, index=False, sheet_name=USER_SHEET)


def add_user(user_name):
    user_name = str(user_name).strip()

    if not user_name:
        return False, "利用者名が空欄です。"

    df = load_users(include_hidden=True)

    if user_name in df["利用者名"].tolist():
        df.loc[df["利用者名"] == user_name, "表示"] = "表示"
        save_users(df)
        return True, f"{user_name} を表示に戻しました。"

    row = {
        "利用者名": user_name,
        "表示": "表示",
    }
    for col in ASSESSMENT_COLUMNS:
        row[col] = ""

    new_row = pd.DataFrame([row])

    df = pd.concat([df, new_row], ignore_index=True)
    save_users(df)

    return True, f"{user_name} を追加しました。"


def hide_user(user_name):
    df = load_users(include_hidden=True)

    if user_name not in df["利用者名"].tolist():
        return False, "対象の利用者が見つかりません。"

    df.loc[df["利用者名"] == user_name, "表示"] = "非表示"
    save_users(df)

    return True, f"{user_name} を入力候補から外しました。過去データは残ります。"


def ensure_data_file():
    ensure_dirs()

    if not DATA_FILE.exists():
        df = pd.DataFrame(columns=COLUMNS)
        df.to_excel(DATA_FILE, index=False, sheet_name=SHEET_NAME)


def load_data():
    ensure_data_file()

    try:
        df = pd.read_excel(DATA_FILE, sheet_name=SHEET_NAME)
    except Exception:
        df = pd.DataFrame(columns=COLUMNS)

    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[COLUMNS]

    if not df.empty:
        df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")

    return df


def save_data(df):
    ensure_dirs()
    df = df.copy()
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNS]

    # 記録日＋利用者名をキーに重複を整理して保存
    if "記録日" in df.columns and "利用者名" in df.columns and not df.empty:
        df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")
        df["利用者名"] = df["利用者名"].astype(str).str.strip()
        df["_検索キー"] = df.apply(
            lambda row: make_record_key(row["記録日"], row["利用者名"]),
            axis=1,
        ) if "make_record_key" in globals() else ""
        if "_検索キー" in df.columns:
            df = df[df["_検索キー"] != ""]
            df = df.drop_duplicates(subset=["_検索キー"], keep="last")
            df = df.drop(columns=["_検索キー"])

    df = df[COLUMNS]
    df.to_excel(DATA_FILE, index=False, sheet_name=SHEET_NAME)


def make_record_key(record_date, user_name):
    """記録日＋利用者名を検索キーとして扱う。"""
    d = pd.to_datetime(record_date, errors="coerce")
    if pd.isna(d):
        return ""
    return f"{d.strftime('%Y-%m-%d')}__{str(user_name).strip()}"


def normalize_key_columns(df):
    """記録日と利用者名を整え、同じ日の同じ利用者が重複しないようにする。"""
    df = df.copy()

    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""

    if not df.empty:
        df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")
        df["利用者名"] = df["利用者名"].astype(str).str.strip()

        df["_検索キー"] = df.apply(
            lambda row: make_record_key(row["記録日"], row["利用者名"]),
            axis=1,
        )

        # 同じ「記録日＋利用者名」が複数ある場合は、最後の記録を正式データとして残す
        df = df[df["_検索キー"] != ""]
        df = df.drop_duplicates(subset=["_検索キー"], keep="last")
        df = df.drop(columns=["_検索キー"])

    return df[COLUMNS]


def find_record_index(df, record_date, user_name):
    """記録日＋利用者名で既存データのindexを探す。
    日付は年月日だけで比較し、時間部分は無視する。
    """
    if df.empty:
        return None

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    work["利用者名"] = work["利用者名"].astype(str).str.strip()

    target_date = pd.to_datetime(record_date, errors="coerce")
    if pd.isna(target_date):
        return None

    target_date_value = target_date.date()
    target_user = str(user_name).strip()

    mask = (
        (work["記録日"].dt.date == target_date_value)
        & (work["利用者名"] == target_user)
    )

    matches = work.index[mask].tolist()

    if not matches:
        return None

    return matches[0]


def upsert_record(record):
    """記録日＋利用者名をキーにして、なければ登録、あれば更新する。"""
    df = load_data()
    df = normalize_key_columns(df)

    idx = find_record_index(
        df,
        record["記録日"],
        record["利用者名"],
    )

    if idx is None:
        new_df = pd.DataFrame([record], columns=COLUMNS)
        df = pd.concat([df, new_df], ignore_index=True)
        action = "登録"
    else:
        for col in COLUMNS:
            df.loc[idx, col] = record.get(col, "")
        action = "更新"

    df = normalize_key_columns(df)
    save_data(df)

    return action


def append_record(record):
    """互換用。内部ではupsert_recordを使う。"""
    return upsert_record(record)


def to_number(series):
    return pd.to_numeric(series, errors="coerce")


def safe_float(value, default):
    try:
        if pd.isna(value) or value == "":
            return default
        return float(value)
    except Exception:
        return default


def safe_int(value, default):
    try:
        if pd.isna(value) or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def safe_text(value):
    if pd.isna(value):
        return ""
    return str(value)



def build_excretion_inputs():
    """時系列で排泄状況を入力し、合計回数と詳細データを返す。"""
    st.subheader("排泄状況")
    st.caption("日中帯（9時〜17時）と夜間帯（18時〜翌8時）を時系列で記録できます。尿量・便量が「なし」の場合、性状は保存時に空欄になります。")

    excretion_data = {}
    urine_count = 0
    stool_count = 0

    def slot_card(slot, time_label, card_color, border_color):
        nonlocal urine_count, stool_count, excretion_data

        st.markdown(
            f"""
            <div style='background:{card_color}; padding:12px; border-radius:14px; border:1px solid {border_color}; margin-bottom:10px;'>
                <b style='font-size:16px;'>{slot}</b><br>
                <span style='font-size:12px; color:#666;'>{time_label}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # 「なし」を選んでも性状欄は選べるようにしておく。
        # 保存時に量が「なし」の場合は性状を空欄保存する。
        urine_amount = st.selectbox(
            f"{slot} 尿量",
            URINE_AMOUNT_OPTIONS,
            index=0,
            key=f"excretion_{slot}_urine_amount",
        )

        urine_type = st.selectbox(
            f"{slot} 尿性状",
            URINE_TYPE_OPTIONS,
            index=0,
            key=f"excretion_{slot}_urine_type",
        )

        stool_amount = st.selectbox(
            f"{slot} 便量",
            STOOL_AMOUNT_OPTIONS,
            index=0,
            key=f"excretion_{slot}_stool_amount",
        )

        stool_type = st.selectbox(
            f"{slot} 便性状",
            STOOL_TYPE_OPTIONS,
            index=0,
            key=f"excretion_{slot}_stool_type",
        )

        excretion_data[f"{slot}尿量"] = urine_amount
        excretion_data[f"{slot}尿性状"] = "" if urine_amount == "なし" else urine_type
        excretion_data[f"{slot}便量"] = stool_amount
        excretion_data[f"{slot}便性状"] = "" if stool_amount == "なし" else stool_type

        if urine_amount != "なし":
            urine_count += 1

        if stool_amount != "なし":
            stool_count += 1

    st.markdown("#### ☀️ 日中帯（9時〜17時）")
    day_cols = st.columns(3)

    for col, slot_info in zip(day_cols, EXCRETION_SLOTS[:3]):
        slot, time_label = slot_info
        with col:
            slot_card(
                slot=slot,
                time_label=time_label,
                card_color="#FFF7EC",
                border_color="#E5D5BF",
            )

    st.markdown("#### 🌙 夜間帯（18時〜翌8時）")
    night_cols = st.columns(3)

    for col, slot_info in zip(night_cols, EXCRETION_SLOTS[3:]):
        slot, time_label = slot_info
        with col:
            slot_card(
                slot=slot,
                time_label=time_label,
                card_color="#EEF4FA",
                border_color="#C9D8E6",
            )

    st.info(f"自動集計：排尿 {urine_count} 回 ／ 排便 {stool_count} 回")

    with st.expander("保存される排泄詳細を確認する"):
        preview = pd.DataFrame([excretion_data])
        st.dataframe(preview, use_container_width=True, hide_index=True)

    return urine_count, stool_count, excretion_data


def build_excretion_admin_summary(df, active_users, target_start, target_end):
    """管理者向け：排泄詳細を集計し、状況把握用データを返す。"""
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")

    work = work[
        (work["記録日"].dt.date >= target_start)
        & (work["記録日"].dt.date <= target_end)
    ].copy()

    if work.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    summary_rows = []

    for user in active_users:
        user_df = work[work["利用者名"] == user].copy()

        if user_df.empty:
            summary_rows.append(
                {
                    "利用者名": user,
                    "記録日数": 0,
                    "排尿回数合計": 0,
                    "排便回数合計": 0,
                    "排便なし日数": 0,
                    "濃縮尿記録": 0,
                    "下痢便記録": 0,
                    "水様便記録": 0,
                    "注意メモ": "記録なし",
                }
            )
            continue

        urine_total = int(to_number(user_df["排尿回数"]).fillna(0).sum())
        stool_total = int(to_number(user_df["排便回数"]).fillna(0).sum())
        no_stool_days = int((to_number(user_df["排便回数"]).fillna(0) == 0).sum())

        concentrated_urine = 0
        diarrhea = 0
        watery = 0

        for slot, _ in EXCRETION_SLOTS:
            urine_type_col = f"{slot}尿性状"
            stool_type_col = f"{slot}便性状"

            if urine_type_col in user_df.columns:
                concentrated_urine += int((user_df[urine_type_col].fillna("") == "濃縮尿").sum())

            if stool_type_col in user_df.columns:
                diarrhea += int((user_df[stool_type_col].fillna("") == "下痢便").sum())
                watery += int((user_df[stool_type_col].fillna("") == "水様便").sum())

        notes = []
        if no_stool_days >= 3:
            notes.append("排便なし日数が多い")
        if concentrated_urine > 0:
            notes.append("濃縮尿あり")
        if diarrhea > 0:
            notes.append("下痢便あり")
        if watery > 0:
            notes.append("水様便あり")

        summary_rows.append(
            {
                "利用者名": user,
                "記録日数": len(user_df),
                "排尿回数合計": urine_total,
                "排便回数合計": stool_total,
                "排便なし日数": no_stool_days,
                "濃縮尿記録": concentrated_urine,
                "下痢便記録": diarrhea,
                "水様便記録": watery,
                "注意メモ": "、".join(notes) if notes else "大きな注意記録なし",
            }
        )

    summary_df = pd.DataFrame(summary_rows)

    detail_cols = [
        "記録日",
        "利用者名",
        "排尿回数",
        "排便回数",
    ]

    for slot, _ in EXCRETION_SLOTS:
        detail_cols.extend(
            [
                f"{slot}尿量",
                f"{slot}尿性状",
                f"{slot}便量",
                f"{slot}便性状",
            ]
        )

    detail_cols = [col for col in detail_cols if col in work.columns]
    detail_df = work[detail_cols].sort_values(["記録日", "利用者名"])

    alert_rows = []

    for _, row in work.iterrows():
        alerts = []

        if safe_int(row.get("排便回数"), 0) == 0:
            alerts.append("排便なし")

        for slot, _ in EXCRETION_SLOTS:
            urine_amount = safe_text(row.get(f"{slot}尿量", ""))
            urine_type = safe_text(row.get(f"{slot}尿性状", ""))
            stool_amount = safe_text(row.get(f"{slot}便量", ""))
            stool_type = safe_text(row.get(f"{slot}便性状", ""))

            if urine_type == "濃縮尿":
                alerts.append(f"{slot}：濃縮尿")
            if stool_type in ["下痢便", "水様便"]:
                alerts.append(f"{slot}：{stool_type}")
            if stool_amount == "大":
                alerts.append(f"{slot}：便量大")
            if urine_amount == "大":
                alerts.append(f"{slot}：尿量大")

        if alerts:
            alert_rows.append(
                {
                    "記録日": row.get("記録日"),
                    "利用者名": row.get("利用者名"),
                    "確認内容": "、".join(alerts),
                    "気になる変化": row.get("気になる変化", ""),
                    "入力者": row.get("入力者", ""),
                }
            )

    alert_df = pd.DataFrame(alert_rows)

    return summary_df, detail_df, alert_df


def build_excretion_summary_text(target):
    """排泄詳細を家族向け・管理者向けに簡潔にまとめる。"""
    if target.empty:
        return ""

    lines = []
    for _, row in target.iterrows():
        date_text = row["記録日"].strftime("%m/%d") if pd.notna(row.get("記録日")) else ""
        slot_notes = []

        for slot, _ in EXCRETION_SLOTS:
            urine_amount = safe_text(row.get(f"{slot}尿量", ""))
            urine_type = safe_text(row.get(f"{slot}尿性状", ""))
            stool_amount = safe_text(row.get(f"{slot}便量", ""))
            stool_type = safe_text(row.get(f"{slot}便性状", ""))

            notes = []
            if urine_amount and urine_amount != "なし":
                notes.append(f"尿{urine_amount}" + (f"・{urine_type}" if urine_type else ""))
            if stool_amount and stool_amount != "なし":
                notes.append(f"便{stool_amount}" + (f"・{stool_type}" if stool_type else ""))

            if notes:
                slot_notes.append(f"{slot}：" + "／".join(notes))

        if slot_notes:
            lines.append(f"{date_text}　" + "、".join(slot_notes))

    return "\n".join(lines[:10])


def get_user_assessment(user_name):
    """利用者マスタからアセスメント情報を取得する。"""
    try:
        df_users = load_users(include_hidden=True)
        row = df_users[df_users["利用者名"] == user_name]

        if row.empty:
            return {}

        row = row.iloc[0]
        return {
            col: safe_text(row.get(col, "")).strip()
            for col in ASSESSMENT_COLUMNS
            if safe_text(row.get(col, "")).strip()
        }
    except Exception:
        return {}


def build_assessment_context_text(user_name):
    """AIレポートや申し送りに使いやすい背景情報を文章化する。"""
    assessment = get_user_assessment(user_name)

    if not assessment:
        return ""

    important_cols = [
        "主訴",
        "生活状況",
        "ADL",
        "IADL",
        "認知機能",
        "健康状態",
        "課題",
        "支援内容",
    ]

    parts = []

    for col in important_cols:
        value = assessment.get(col, "")
        if value:
            parts.append(f"{col}：{value}")

    return "\n".join(parts)


def build_assessment_report_text(user_name):
    """家族向け・PDF向けに、アセスメント情報をやわらかく要約する。"""
    assessment = get_user_assessment(user_name)

    if not assessment:
        return "現在、アセスメント情報は未登録です。今後の記録とあわせて、生活全体の様子を確認していきます。"

    lines = []

    complaint = assessment.get("主訴", "")
    life = assessment.get("生活状況", "")
    adl = assessment.get("ADL", "")
    iadl = assessment.get("IADL", "")
    cognitive = assessment.get("認知機能", "")
    health = assessment.get("健康状態", "")
    issue = assessment.get("課題", "")
    support = assessment.get("支援内容", "")

    if complaint:
        lines.append("ご本人・ご家族のご希望や困りごととして、" + complaint)

    if life:
        lines.append("日々の生活状況として、" + life)

    care_points = []
    if adl:
        care_points.append("ADL：" + adl)
    if iadl:
        care_points.append("IADL：" + iadl)
    if cognitive:
        care_points.append("認知機能：" + cognitive)
    if health:
        care_points.append("健康状態：" + health)

    if care_points:
        lines.append("見守りの視点として、" + "／".join(care_points))

    if issue:
        lines.append("支援上の課題として、" + issue)

    if support:
        lines.append("現在の支援内容として、" + support)

    if not lines:
        return "アセスメント情報は登録されていますが、家族向けに表示する内容はまだ整理中です。"

    lines.append("これらの背景情報をふまえ、日々の健康記録・食事・排泄・ご様子をあわせて確認しています。")

    return "\n\n".join(lines)


def build_admin_assessment_analysis(user_name, target):
    """管理者支援用：アセスメントと月間記録を組み合わせた確認視点を作成する。"""
    assessment_text = build_assessment_context_text(user_name)

    if target.empty:
        record_text = "対象月の健康チェック記録はありません。"
    else:
        low_meal_count = 0
        for col in ["朝食摂取率", "昼食摂取率", "夕食摂取率"]:
            if col in target.columns:
                low_meal_count += int((to_number(target[col]) <= 50).sum())

        no_stool_days = 0
        if "排便回数" in target.columns:
            no_stool_days = int((to_number(target["排便回数"]).fillna(0) == 0).sum())

        memo_count = int((target["気になる変化"].fillna("").astype(str).str.strip() != "").sum())

        record_text = (
            f"対象月の記録件数：{len(target)}件\n"
            f"食事摂取率50％以下の記録：{low_meal_count}件\n"
            f"排便0回の日数：{no_stool_days}日\n"
            f"気になる変化の記録：{memo_count}件"
        )

    return (
        "【アセスメント情報】\n"
        + (assessment_text if assessment_text else "未登録")
        + "\n\n【月間記録から見える確認点】\n"
        + record_text
        + "\n\n【管理者確認の視点】\n"
        "・アセスメント上の課題と、実際の記録にずれがないか確認してください。\n"
        "・食事摂取率、排便状況、気になる変化が、ADL・認知機能・健康状態と関係していないか確認してください。\n"
        "・医療判断ではなく、職員間の共有と見守り方針の整理に使ってください。"
    )


def get_month_data(df, user_name, year, month):
    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    return work[
        (work["利用者名"] == user_name)
        & (work["記録日"].dt.year == int(year))
        & (work["記録日"].dt.month == int(month))
    ].sort_values("記録日")


def create_family_summary_text(target, user_name, year, month):
    if target.empty:
        return (
            f"{user_name}の{year}年{month}月分の記録は、現時点では登録されていません。"
            "今後の記録をもとに、ご様子を継続して確認していきます。"
        )

    lines = []
    record_count = len(target)
    lines.append(
        f"{user_name}の{year}年{month}月の記録は、{record_count}件確認されています。"
        "この文章は医療的な判断ではなく、日々の健康チェック記録をもとにした共有です。"
    )

    assessment_report_text = build_assessment_report_text(user_name)
    if assessment_report_text:
        lines.append(
            "アセスメント情報もふまえた見守りの視点です。\n"
            + assessment_report_text
        )

    temp_mean = to_number(target["体温"]).mean()
    spo2_mean = to_number(target["SpO2"]).mean()
    weight_mean = to_number(target["体重"]).mean()

    health_parts = []
    if not pd.isna(temp_mean):
        health_parts.append(f"体温は平均{round(float(temp_mean), 1)}℃")
    if not pd.isna(spo2_mean):
        health_parts.append(f"SpO2は平均{round(float(spo2_mean), 1)}％")
    if not pd.isna(weight_mean):
        health_parts.append(f"体重は平均{round(float(weight_mean), 1)}kg")


    breakfast_mean = to_number(target["朝食摂取率"]).mean()
    lunch_mean = to_number(target["昼食摂取率"]).mean()
    dinner_mean = to_number(target["夕食摂取率"]).mean()

    urine_mean = to_number(target["排尿回数"]).mean()
    stool_mean = to_number(target["排便回数"]).mean()

    if not pd.isna(breakfast_mean):
        lines.append(
            f"食事摂取率は、朝食平均{round(float(breakfast_mean),1)}％、"
            f"昼食平均{round(float(lunch_mean),1)}％、"
            f"夕食平均{round(float(dinner_mean),1)}％でした。"
        )

    if not pd.isna(urine_mean):
        lines.append(
            f"排泄状況は、排尿回数平均{round(float(urine_mean),1)}回、"
            f"排便回数平均{round(float(stool_mean),1)}回として記録されています。"
        )

    excretion_detail = build_excretion_summary_text(target)
    if excretion_detail:
        lines.append(
            "排泄の詳細記録として、以下の内容が確認されています。\n"
            + excretion_detail
        )

    if health_parts:
        lines.append(
            "記録上、" + "、".join(health_parts) + "として確認されています。"
            "数値は日々の状態を振り返るための目安として扱っています。"
        )

    alerts = []
    temp_alerts = target[to_number(target["体温"]) >= 37.5]
    spo2_alerts = target[to_number(target["SpO2"]) <= 93]
    bp_alerts = target[to_number(target["血圧上"]) >= 160]

    if not temp_alerts.empty:
        dates = "、".join(temp_alerts["記録日"].dt.strftime("%m/%d").tolist()[:5])
        alerts.append(f"{dates}に体温が37.5℃以上の記録")
    if not spo2_alerts.empty:
        dates = "、".join(spo2_alerts["記録日"].dt.strftime("%m/%d").tolist()[:5])
        alerts.append(f"{dates}にSpO2が93％以下の記録")
    if not bp_alerts.empty:
        dates = "、".join(bp_alerts["記録日"].dt.strftime("%m/%d").tolist()[:5])
        alerts.append(f"{dates}に血圧上が160以上の記録")

    if alerts:
        lines.append(
            "今月は、" + "、".join(alerts) + "がありました。"
            "一時的な変動の可能性もあるため、引き続き経過を見ながら確認していきます。"
        )
    else:
        lines.append(
            "記録上、設定した注意目安に該当する大きな変化は目立っていません。"
            "今後も日々の様子を継続して確認していきます。"
        )

    memo_rows = target[target["家族共有メモ"].fillna("").astype(str).str.strip() != ""]
    change_rows = target[target["気になる変化"].fillna("").astype(str).str.strip() != ""]

    if not memo_rows.empty:
        first = memo_rows.iloc[0]
        lines.append(
            f"ご様子として、{first['記録日'].strftime('%m/%d')}の記録に"
            f"「{str(first['家族共有メモ'])[:80]}」とあります。"
            "日々の関わりの中で、ご本人の様子を確認しています。"
        )

    if not change_rows.empty:
        first = change_rows.iloc[0]
        lines.append(
            f"また、{first['記録日'].strftime('%m/%d')}に"
            f"「{str(first['気になる変化'])[:80]}」という記録があります。"
            "必要に応じて職員間で共有しながら見守っています。"
        )

    lines.append(
        "今後も、数値だけでなく表情や生活の様子も含めて、安心して過ごせるよう見守ってまいります。"
    )

    return "\n\n".join(lines)


def create_handover_text(df, target_date):
    if df.empty:
        return "本日の記録はまだありません。"

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    day_df = work[work["記録日"].dt.date == target_date]

    if day_df.empty:
        return "指定日の記録はありません。"

    lines = [
        f"{target_date.strftime('%Y/%m/%d')}の申し送りまとめです。",
        "医療的な判断ではなく、記録内容をもとにした共有用メモです。",
        "",
    ]

    for _, row in day_df.iterrows():
        notes = []
        if str(row.get("気になる変化", "")).strip() and not pd.isna(row.get("気になる変化")):
            notes.append(f"気になる変化：{row.get('気になる変化')}")
        if str(row.get("家族共有メモ", "")).strip() and not pd.isna(row.get("家族共有メモ")):
            notes.append(f"家族共有メモ：{row.get('家族共有メモ')}")

        vital_alerts = []
        if safe_float(row.get("体温"), 0) >= 37.5:
            vital_alerts.append("体温高め")
        if safe_int(row.get("SpO2"), 100) <= 93:
            vital_alerts.append("SpO2低め")
        if safe_int(row.get("血圧上"), 0) >= 160:
            vital_alerts.append("血圧上高め")

        excretion_alerts = []
        for slot, _ in EXCRETION_SLOTS:
            urine_type = safe_text(row.get(f"{slot}尿性状", ""))
            stool_type = safe_text(row.get(f"{slot}便性状", ""))

            if urine_type == "濃縮尿":
                excretion_alerts.append(f"{slot}に濃縮尿")
            if stool_type in ["下痢便", "水様便"]:
                excretion_alerts.append(f"{slot}に{stool_type}")

        if excretion_alerts:
            notes.append("排泄確認：" + "、".join(excretion_alerts[:6]))

        if vital_alerts:
            notes.append("確認目安：" + "、".join(vital_alerts))

        if notes:
            lines.append(f"■ {row.get('利用者名')}")
            lines.extend([f"・{x}" for x in notes])
            lines.append("")

    if len(lines) <= 3:
        lines.append("記録上、特に申し送り対象となるメモや注意目安はありません。")

    lines.append("引き続き、普段との違いがないかを確認しながら見守ります。")
    return "\n".join(lines)


def get_today_dashboard(df, active_users):
    today = pd.Timestamp(date.today())

    if df.empty:
        today_df = pd.DataFrame(columns=COLUMNS)
    else:
        work = df.copy()
        work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
        today_df = work[work["記録日"].dt.date == today.date()].copy()

    entered_users = today_df["利用者名"].dropna().astype(str).unique().tolist()
    missing_users = [user for user in active_users if user not in entered_users]

    temp_alert = today_df[to_number(today_df["体温"]) >= 37.5] if not today_df.empty else today_df
    spo2_alert = today_df[to_number(today_df["SpO2"]) <= 93] if not today_df.empty else today_df
    bp_alert = today_df[to_number(today_df["血圧上"]) >= 160] if not today_df.empty else today_df

    alert_rows = pd.concat([temp_alert, spo2_alert, bp_alert]).drop_duplicates() if not today_df.empty else today_df

    memo_count = 0
    if not today_df.empty:
        memo_count = int(
            (
                today_df["家族共有メモ"].fillna("").astype(str).str.strip() != ""
            ).sum()
        )

    return {
        "today_df": today_df,
        "entered_users": entered_users,
        "missing_users": missing_users,
        "alert_rows": alert_rows,
        "memo_count": memo_count,
    }


def show_dashboard(active_users):
    st.header("管理者ダッシュボード")
    st.caption("今日の入力状況・未入力・注意記録を一画面で確認できます。")

    df = load_data()
    dashboard = get_today_dashboard(df, active_users)

    today_df = dashboard["today_df"]
    entered_count = len(dashboard["entered_users"])
    total_count = len(active_users)
    missing_users = dashboard["missing_users"]
    alert_rows = dashboard["alert_rows"]
    memo_count = dashboard["memo_count"]

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("本日の入力済み", f"{entered_count} / {total_count} 名")

    with col2:
        st.metric("未入力", f"{len(missing_users)} 名")

    with col3:
        st.metric("注意記録", f"{len(alert_rows)} 件")

    with col4:
        st.metric("家族共有メモあり", f"{memo_count} 件")

    st.divider()

    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader("未入力の利用者様")
        if missing_users:
            st.warning("、".join(missing_users))
        else:
            st.success("本日の入力は全員分そろっています。")

    with col_right:
        st.subheader("本日の記録件数")
        if today_df.empty:
            st.info("本日の記録はまだありません。")
        else:
            count_df = today_df.groupby("利用者名").size().reset_index(name="記録件数")
            st.dataframe(count_df, use_container_width=True, hide_index=True)

    st.subheader("注意記録の確認")
    st.caption("目安：体温37.5℃以上、SpO2 93％以下、血圧上160以上。医療判断ではなく、見落とし防止のための確認欄です。")


    st.divider()
    st.subheader("排便アラート")

    constipation_users = []

    for user in active_users:

        user_df = df[df["利用者名"] == user].copy()

        if user_df.empty:
            continue

        user_df = user_df.sort_values("記録日")
        recent = user_df.tail(3)

        if len(recent) < 3:
            continue

        stool_values = pd.to_numeric(
            recent["排便回数"],
            errors="coerce"
        ).fillna(0)

        if (stool_values == 0).all():
            constipation_users.append(user)

    if constipation_users:
        st.warning(
            "未排便が3日続いています：" +
            "、".join(constipation_users)
        )
    else:
        st.success("未排便3日アラートはありません。")

    if alert_rows.empty:
        st.success("本日、設定した注意目安に該当する記録はありません。")
    else:
        show_cols = [
            "記録日",
            "利用者名",
            "体温",
            "血圧上",
            "血圧下",
            "脈拍",
            "SpO2",
            "体重",
            "気になる変化",
            "入力者",
        ]
        st.dataframe(alert_rows[show_cols], use_container_width=True, hide_index=True)


def create_family_report(df, user_name, year, month):
    target = get_month_data(df, user_name, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "家族向けレポート"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    title_fill = PatternFill("solid", fgColor="EADFCB")
    section_fill = PatternFill("solid", fgColor="DDEFE2")
    note_fill = PatternFill("solid", fgColor="FFF8E7")
    summary_fill = PatternFill("solid", fgColor="F4F0E8")

    ws.merge_cells("A1:H1")
    ws["A1"] = "ご家族向け 健康・生活レポート"
    ws["A1"].font = Font(name="Meiryo", size=16, bold=True)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "利用者名"
    ws["B3"] = user_name
    ws["D3"] = "対象月"
    ws["E3"] = f"{year}年{month}月"

    ws.merge_cells("A5:H5")
    ws["A5"] = "今月の健康状態"
    ws["A5"].font = Font(name="Meiryo", bold=True)
    ws["A5"].fill = section_fill

    metrics = [
        ("体温平均", "体温", "℃"),
        ("血圧上平均", "血圧上", ""),
        ("血圧下平均", "血圧下", ""),
        ("脈拍平均", "脈拍", "回/分"),
        ("SpO2平均", "SpO2", "%"),
        ("体重平均", "体重", "kg"),
    ]

    start_row = 7

    for i, (label, col, unit) in enumerate(metrics):
        row = start_row + i
        ws[f"A{row}"] = label

        if target.empty:
            value = ""
        else:
            value = to_number(target[col]).mean()
            value = "" if pd.isna(value) else round(float(value), 1)

        ws[f"B{row}"] = value
        ws[f"C{row}"] = unit

    ws.merge_cells("A14:H14")
    ws["A14"] = "今月のまとめ"
    ws["A14"].font = Font(name="Meiryo", bold=True)
    ws["A14"].fill = section_fill

    summary_text = create_family_summary_text(target, user_name, year, month)
    ws.merge_cells("A15:H20")
    ws["A15"] = summary_text
    ws["A15"].fill = summary_fill
    ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A22:H22")
    ws["A22"] = "アセスメントに基づく見守りの視点"
    ws["A22"].font = Font(name="Meiryo", bold=True)
    ws["A22"].fill = section_fill

    assessment_report_text = build_assessment_report_text(user_name)
    ws.merge_cells("A23:H27")
    ws["A23"] = assessment_report_text
    ws["A23"].fill = summary_fill
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="top")

    row = 29
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).value = "日付別のご様子（家族共有メモ）"
    ws.cell(row=row, column=1).font = Font(name="Meiryo", bold=True)
    ws.cell(row=row, column=1).fill = section_fill

    row += 1
    ws.cell(row=row, column=1).value = "日付"
    ws.cell(row=row, column=2).value = "家族共有メモ"
    ws.cell(row=row, column=1).fill = note_fill
    ws.cell(row=row, column=2).fill = note_fill

    row += 1
    memo_rows = target[target["家族共有メモ"].fillna("").astype(str).str.strip() != ""]

    if memo_rows.empty:
        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=2).value = "記録された家族共有メモはありません。"
        row += 1
    else:
        for _, rec in memo_rows.iterrows():
            ws.cell(row=row, column=1).value = rec["記録日"].strftime("%m/%d")
            ws.cell(row=row, column=2).value = str(rec["家族共有メモ"])
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1).value = "日付別の気になる変化"
    ws.cell(row=row, column=1).font = Font(name="Meiryo", bold=True)
    ws.cell(row=row, column=1).fill = section_fill

    row += 1
    ws.cell(row=row, column=1).value = "日付"
    ws.cell(row=row, column=2).value = "気になる変化"
    ws.cell(row=row, column=1).fill = note_fill
    ws.cell(row=row, column=2).fill = note_fill

    row += 1
    change_rows = target[target["気になる変化"].fillna("").astype(str).str.strip() != ""]

    if change_rows.empty:
        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=2).value = "記録された気になる変化はありません。"
        row += 1
    else:
        for _, rec in change_rows.iterrows():
            ws.cell(row=row, column=1).value = rec["記録日"].strftime("%m/%d")
            ws.cell(row=row, column=2).value = str(rec["気になる変化"])
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=8)
    ws.cell(row=row, column=1).value = (
        "※このレポートは、施設内の健康チェック記録をもとにした共有資料です。"
        "医療的な診断・治療効果の判断を行うものではありません。"
    )
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.column_dimensions["B"].width = 70

    for cells in ws.iter_rows():
        for cell in cells:
            cell.font = Font(
                name="Meiryo",
                size=10,
                bold=cell.font.bold if cell.font else False,
            )
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    file_name = f"家族向けレポート_{user_name}_{year}年{month}月.xlsx"
    report_path = REPORT_DIR / file_name
    wb.save(report_path)

    return report_path, target, summary_text




# =========================
# ひだまりレポートPDF作成
# 線画・シンプル・モダン版
# =========================
def create_chart_image(target, item):
    """PDFに差し込むバイタル推移グラフ画像を作成する。線画風のシンプルデザイン。"""
    if target.empty or item not in target.columns:
        return None

    chart_df = target[["記録日", item]].copy()
    chart_df["記録日"] = pd.to_datetime(chart_df["記録日"], errors="coerce")
    chart_df[item] = pd.to_numeric(chart_df[item], errors="coerce")
    chart_df = chart_df.dropna().sort_values("記録日")

    if chart_df.empty:
        return None

    import matplotlib.pyplot as plt
    from io import BytesIO

    fig, ax = plt.subplots(figsize=(6.2, 2.15))

    # モダンな線画風：色を抑え、線と余白を優先
    ax.plot(
        chart_df["記録日"],
        chart_df[item],
        marker="o",
        linewidth=1.8,
        markersize=4,
        color="#2F3437",
        markerfacecolor="white",
        markeredgecolor="#2F3437",
    )

    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.grid(True, color="#E6E6E6", linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#BBBBBB")
    ax.spines["bottom"].set_color("#BBBBBB")
    ax.tick_params(axis="both", colors="#555555", labelsize=8)
    fig.autofmt_xdate(rotation=30)

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def create_hidamari_report_pdf(df, user_name, year, month):
    """ご家族向けのPDF『ひだまりレポート』を作成する。線画・シンプル・モダン版。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image,
        PageBreak,
        KeepTogether,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.graphics.shapes import Drawing, Circle, Rect, Polygon, Line, String, Path

    ensure_dirs()

    # 日本語フォント登録
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))

    target = get_month_data(df, user_name, year, month)
    summary_text = create_family_summary_text(target, user_name, year, month)

    file_name = f"ひだまりレポート_{user_name}_{year}年{month}月.pdf"
    pdf_path = REPORT_DIR / file_name

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        rightMargin=17 * mm,
        leftMargin=17 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    # 色：線画ベース。黒・グレー・淡いベージュのみ
    ink = colors.HexColor("#2F3437")
    gray = colors.HexColor("#7A7A7A")
    light_gray = colors.HexColor("#D9D9D9")
    pale = colors.HexColor("#F7F4EE")
    pale2 = colors.HexColor("#FBFAF7")
    accent = colors.HexColor("#6F7F72")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "jp_title_modern",
        parent=styles["Title"],
        fontName="HeiseiKakuGo-W5",
        fontSize=24,
        leading=30,
        alignment=1,
        textColor=ink,
        spaceAfter=3,
    )
    # 温かみのあるモダンデザイン
    subtitle_style = ParagraphStyle(
        "jp_subtitle_modern",
        parent=styles["Normal"],
        fontName="HeiseiMin-W3",
        fontSize=9.8,
        leading=15,
        alignment=1,
        textColor=gray,
        spaceAfter=10,
    )
    h2_style = ParagraphStyle(
        "jp_h2_modern",
        parent=styles["Heading2"],
        fontName="HeiseiKakuGo-W5",
        fontSize=13.2,
        leading=18,
        textColor=ink,
        spaceBefore=8,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "jp_body_modern",
        parent=styles["BodyText"],
        fontName="HeiseiMin-W3",
        fontSize=10.2,
        leading=17,
        textColor=colors.HexColor("#333333"),
    )
    small_style = ParagraphStyle(
        "jp_small_modern",
        parent=styles["BodyText"],
        fontName="HeiseiMin-W3",
        fontSize=8.5,
        leading=12,
        textColor=gray,
    )
    label_style = ParagraphStyle(
        "jp_label_modern",
        parent=styles["BodyText"],
        fontName="HeiseiKakuGo-W5",
        fontSize=9,
        leading=12,
        textColor=ink,
    )

    def line_art_header():
        """季節感のある、引き算ベースの温かい線画ヘッダー。"""
        d = Drawing(500, 96)

        # 背景の淡い帯
        d.add(Rect(0, 0, 500, 96,
                   fillColor=colors.HexColor("#FCFAF5"),
                   strokeColor=None))

        # 上下ライン
        d.add(Line(0, 84, 500, 84,
                   strokeColor=colors.HexColor("#DDD7CC"),
                   strokeWidth=0.8))

        d.add(Line(0, 10, 500, 10,
                   strokeColor=colors.HexColor("#DDD7CC"),
                   strokeWidth=0.8))

        # 季節の太陽
        d.add(Circle(
            62, 50, 16,
            fillColor=colors.HexColor("#FFF6E8"),
            strokeColor=colors.HexColor("#B68A5A"),
            strokeWidth=0.9
        ))

        for x1, y1, x2, y2 in [
            (62, 76, 62, 68),
            (62, 31, 62, 23),
            (37, 50, 46, 50),
            (78, 50, 88, 50),
            (45, 67, 51, 61),
            (79, 67, 73, 61),
            (45, 33, 51, 39),
            (79, 33, 73, 39),
        ]:
            d.add(Line(
                x1, y1, x2, y2,
                strokeColor=colors.HexColor("#B68A5A"),
                strokeWidth=0.8
            ))

        # 家
        d.add(Polygon(
            [382, 36, 425, 67, 468, 36],
            fillColor=colors.HexColor("#FAF7F1"),
            strokeColor=colors.HexColor("#5C5C5C"),
            strokeWidth=1.0
        ))

        d.add(Rect(
            394, 23, 62, 33,
            fillColor=colors.white,
            strokeColor=colors.HexColor("#5C5C5C"),
            strokeWidth=1.0
        ))

        d.add(Rect(
            420, 23, 12, 21,
            fillColor=colors.HexColor("#F7F1E7"),
            strokeColor=colors.HexColor("#5C5C5C"),
            strokeWidth=0.8
        ))

        # 季節の植物
        flower_colors = [
            "#D8A48F",
            "#A7BFA3",
            "#E7C9A9",
            "#CBB8A9",
        ]

        flower_positions = [(150, 24), (170, 30), (192, 24), (212, 29)]

        for i, (x, y) in enumerate(flower_positions):

            d.add(Line(
                x, 18, x, y + 12,
                strokeColor=colors.HexColor("#8A9A7B"),
                strokeWidth=0.7
            ))

            d.add(Circle(
                x - 4, y + 8, 3.8,
                fillColor=colors.HexColor(flower_colors[i % len(flower_colors)]),
                strokeColor=colors.HexColor("#7A7A7A"),
                strokeWidth=0.3
            ))

            d.add(Circle(
                x + 4, y + 8, 3.8,
                fillColor=colors.HexColor(flower_colors[(i+1) % len(flower_colors)]),
                strokeColor=colors.HexColor("#7A7A7A"),
                strokeWidth=0.3
            ))

            d.add(Circle(
                x, y + 14, 3.8,
                fillColor=colors.HexColor(flower_colors[(i+2) % len(flower_colors)]),
                strokeColor=colors.HexColor("#7A7A7A"),
                strokeWidth=0.3
            ))

        # ゆるやかな道
        path = Path()
        path.moveTo(240, 18)
        path.curveTo(260, 32, 285, 42, 320, 48)
        d.add(path)
        d.contents[-1].strokeColor = colors.HexColor("#D5CEC3")
        d.contents[-1].strokeWidth = 1.0
        d.contents[-1].fillColor = None

        return d

    def section_line(title):
        """見出しを線画風に整える。"""
        d = Drawing(500, 20)
        d.add(Line(0, 8, 500, 8, strokeColor=light_gray, strokeWidth=0.8))
        d.add(Circle(5, 8, 3, fillColor=None, strokeColor=ink, strokeWidth=0.8))
        return KeepTogether([Paragraph(title, h2_style), d])

    def make_note_box(text):
        box = Table(
            [[Paragraph(text.replace("\n", "<br/>"), body_style)]],
            colWidths=[168 * mm],
        )
        box.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), pale2),
                    ("BOX", (0, 0), (-1, -1), 0.6, light_gray),
                    ("LEFTPADDING", (0, 0), (-1, -1), 9),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return box

    story = []

    # 表紙
    story.append(line_art_header())
    story.append(Spacer(1, 6))
    story.append(Paragraph("ひだまりレポート", title_style))
    story.append(Paragraph("ご家族へお届けする、月間の健康・生活記録", subtitle_style))
    story.append(Spacer(1, 6))

    info_table = Table(
        [
            ["利用者様", str(user_name), "対象月", f"{year}年{month}月"],
            ["作成日", date.today().strftime("%Y/%m/%d"), "作成元", "ひだまり弐番館"],
        ],
        colWidths=[28 * mm, 55 * mm, 28 * mm, 55 * mm],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "HeiseiKakuGo-W5"),
                ("FONTSIZE", (0, 0), (-1, -1), 9.2),
                ("TEXTCOLOR", (0, 0), (-1, -1), ink),
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.65, light_gray),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, light_gray),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEADING", (0, 0), (-1, -1), 13),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(info_table)
    story.append(Spacer(1, 12))

    story.append(section_line("今月のまとめ"))
    for paragraph in summary_text.split("\n\n"):
        story.append(make_note_box(paragraph))
        story.append(Spacer(1, 5))

    story.append(Spacer(1, 8))
    story.append(section_line("アセスメントに基づく見守りの視点"))
    assessment_report_text = build_assessment_report_text(user_name)
    story.append(make_note_box(assessment_report_text))
    story.append(Spacer(1, 8))

    story.append(section_line("健康状態の目安"))

    metrics = [
        ("体温平均", "体温", "℃"),
        ("血圧上平均", "血圧上", ""),
        ("血圧下平均", "血圧下", ""),
        ("脈拍平均", "脈拍", "回/分"),
        ("SpO2平均", "SpO2", "%"),
        ("体重平均", "体重", "kg"),
    ]

    metric_rows = [["項目", "記録上の平均", "単位"]]
    for label, col, unit in metrics:
        if target.empty:
            value = ""
        else:
            value = to_number(target[col]).mean()
            value = "" if pd.isna(value) else round(float(value), 1)
        metric_rows.append([label, value, unit])

    metric_table = Table(metric_rows, colWidths=[58 * mm, 48 * mm, 32 * mm])
    metric_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "HeiseiKakuGo-W5"),
                ("FONTSIZE", (0, 0), (-1, -1), 9.3),
                ("TEXTCOLOR", (0, 0), (-1, -1), ink),
                ("BACKGROUND", (0, 0), (-1, 0), pale),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.65, light_gray),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, light_gray),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(metric_table)

    story.append(PageBreak())

    # 2ページ目
    story.append(line_art_header())
    story.append(Paragraph("バイタル推移グラフ", h2_style))
    story.append(Paragraph("日々の記録をもとに、体調の経過を見える化しています。", body_style))
    story.append(Spacer(1, 6))

    chart_items = ["体温", "血圧上", "SpO2", "体重"]
    chart_buffers = []
    for item in chart_items:
        buf = create_chart_image(target, item)
        if buf is not None:
            chart_buffers.append((item, buf))

    if not chart_buffers:
        story.append(Paragraph("グラフ化できる記録はまだありません。", body_style))
    else:
        chart_table_rows = []
        for i in range(0, len(chart_buffers), 2):
            cells = []
            for item, buf in chart_buffers[i:i + 2]:
                block = [
                    Paragraph(item, label_style),
                    Image(buf, width=78 * mm, height=34 * mm),
                ]
                cells.append(block)
            if len(cells) == 1:
                cells.append("")
            chart_table_rows.append(cells)

        chart_table = Table(chart_table_rows, colWidths=[83 * mm, 83 * mm])
        chart_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOX", (0, 0), (-1, -1), 0.45, light_gray),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, light_gray),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(chart_table)

    story.append(PageBreak())

    # 3ページ目
    story.append(line_art_header())
    story.append(Paragraph("日々のご様子", h2_style))

    memo_rows = target[target["家族共有メモ"].fillna("").astype(str).str.strip() != ""] if not target.empty else pd.DataFrame()
    change_rows = target[target["気になる変化"].fillna("").astype(str).str.strip() != ""] if not target.empty else pd.DataFrame()

    def make_record_table(rows_df, value_col, empty_text):
        if rows_df.empty:
            return make_note_box(empty_text)

        rows = [["日付", "内容"]]
        for _, rec in rows_df.iterrows():
            rows.append(
                [
                    rec["記録日"].strftime("%m/%d"),
                    Paragraph(str(rec[value_col]).replace("\n", "<br/>"), body_style),
                ]
            )

        table = Table(rows, colWidths=[24 * mm, 140 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "HeiseiKakuGo-W5"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("TEXTCOLOR", (0, 0), (-1, -1), ink),
                    ("BACKGROUND", (0, 0), (-1, 0), pale),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.65, light_gray),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, light_gray),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        return table

    story.append(section_line("家族共有メモ"))
    story.append(make_record_table(memo_rows, "家族共有メモ", "記録された家族共有メモはありません。"))
    story.append(Spacer(1, 10))

    story.append(section_line("気になる変化"))
    story.append(make_record_table(change_rows, "気になる変化", "記録された気になる変化はありません。"))

    story.append(Spacer(1, 14))

    footer_drawing = Drawing(500, 32)
    footer_drawing.add(Line(0, 22, 500, 22, strokeColor=light_gray, strokeWidth=0.7))
    footer_drawing.add(Circle(20, 10, 5, fillColor=None, strokeColor=accent, strokeWidth=0.8))
    footer_drawing.add(Line(25, 10, 65, 10, strokeColor=accent, strokeWidth=0.8))
    story.append(footer_drawing)

    story.append(
        Paragraph(
            "※このレポートは、施設内の健康チェック記録をもとにした共有資料です。医療的な診断・治療効果の判断を行うものではありません。",
            small_style,
        )
    )

    doc.build(story)
    return pdf_path, summary_text, target



# =========================
# アプリ本体
# =========================
logout_button()

st.title("健康チェックWebアプリ")
st.caption("管理者支援・職員入力・家族共有を一体化した健康チェックシステムです。")

ensure_data_file()
ensure_user_file()


# =========================
# サイドバーメニュー
# =========================
if st.session_state.role == "admin":
    st.sidebar.success("管理者モード")
    menu_items = [
        "管理者ダッシュボード",
        "健康チェック入力",
        "過去データ管理",
        "入力データ確認",
        "家族向けレポート作成",
        "ひだまりレポートPDF",
        "管理者支援",
        "排泄詳細管理",
        "利用者マスタ管理",
    ]
else:
    st.sidebar.info("職員モード")
    menu_items = [
        "健康チェック入力",
        "過去データ管理",
    ]

menu = st.sidebar.radio("メニュー", menu_items)


active_users = load_users(include_hidden=False)["利用者名"].tolist()
all_users_df = load_users(include_hidden=True)
all_users = all_users_df["利用者名"].tolist()


if not active_users:
    st.warning("表示中の利用者がいません。管理者に利用者マスタの確認を依頼してください。")

if st.session_state.role == "staff":
    st.markdown(
        """
        <div class='info-box'>
        <b>お疲れ様です。</b><br>
        今日の健康チェック入力をお願いします。小さな変化も、利用者様の安心につながります。
        </div>
        """,
        unsafe_allow_html=True,
    )
elif st.session_state.role == "admin":
    st.markdown(
        """
        <div class='info-box'>
        <b>管理者モードです。</b><br>
        入力状況、注意記録、家族レポート、申し送り支援を確認できます。
        </div>
        """,
        unsafe_allow_html=True,
    )


# =========================
# 管理者ダッシュボード
# =========================
if menu == "管理者ダッシュボード":
    show_dashboard(active_users)


# =========================
# 健康チェック入力
# =========================
elif menu == "健康チェック入力":
    st.header("健康チェック入力")

    if st.session_state.role == "staff":
        st.markdown("### お疲れ様です。")
        st.write("利用者様の今日の健康状態を、時系列でわかりやすく入力してください。")

    if not active_users:
        st.stop()

    with st.form("health_form", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)

        with col1:
            record_date = st.date_input("記録日", value=date.today())

        with col2:
            user_name = st.selectbox("利用者名", active_users)

        with col3:
            input_staff = st.text_input("入力者", placeholder="例：藤野")

        existing_df = load_data()
        existing_idx = find_record_index(existing_df, record_date, user_name)

        if existing_idx is None:
            st.markdown(
                """
                <div style='background:#EAF4FF; border:1px solid #9CC7F0; color:#174A7C; padding:12px 14px; border-radius:10px; margin:8px 0 12px 0;'>
                    <b>この記録日・利用者名のデータはありません。</b><br>
                    登録すると新規データとして保存されます。
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                """
                <div style='background:#FFF3E0; border:1px solid #F0B36A; color:#8A4B00; padding:12px 14px; border-radius:10px; margin:8px 0 12px 0;'>
                    <b>この記録日・利用者名のデータは既にあります。</b><br>
                    登録すると上書き更新されます。
                </div>
                """,
                unsafe_allow_html=True,
            )

        st.divider()
        st.subheader("バイタル")

        col4, col5, col6 = st.columns(3)

        with col4:
            temp = st.number_input(
                "体温",
                min_value=30.0,
                max_value=45.0,
                value=36.5,
                step=0.1,
            )

        with col5:
            bp_high = st.number_input(
                "血圧上",
                min_value=50,
                max_value=250,
                value=120,
                step=1,
            )

        with col6:
            bp_low = st.number_input(
                "血圧下",
                min_value=30,
                max_value=150,
                value=75,
                step=1,
            )

        col7, col8, col9 = st.columns(3)

        with col7:
            pulse = st.number_input(
                "脈拍",
                min_value=30,
                max_value=200,
                value=70,
                step=1,
            )

        with col8:
            spo2 = st.number_input(
                "SpO2",
                min_value=70,
                max_value=100,
                value=96,
                step=1,
            )

        with col9:
            weight = st.number_input(
                "体重",
                min_value=0.0,
                max_value=200.0,
                value=50.0,
                step=0.1,
            )

        st.divider()
        st.subheader("食事摂取率")

        meal1, meal2, meal3 = st.columns(3)

        with meal1:
            breakfast = st.slider("朝食", 0, 100, 80, step=10)

        with meal2:
            lunch = st.slider("昼食", 0, 100, 80, step=10)

        with meal3:
            dinner = st.slider("夕食", 0, 100, 80, step=10)

        st.divider()
        urine_count, stool_count, excretion_data = build_excretion_inputs()

        st.divider()
        family_memo = st.text_area(
            "家族共有メモ",
            placeholder="ご家族へ共有してよい内容を入力",
        )

        changes = st.text_area(
            "気になる変化",
            placeholder="食事、睡眠、歩行、表情、体調、排泄状況など",
        )

        submitted = st.form_submit_button("登録する")

    if submitted:
        record = {
            "記録日": record_date,
            "利用者名": user_name,
            "体温": temp,
            "血圧上": bp_high,
            "血圧下": bp_low,
            "脈拍": pulse,
            "SpO2": spo2,
            "体重": weight,
            "朝食摂取率": breakfast,
            "昼食摂取率": lunch,
            "夕食摂取率": dinner,
            "排尿回数": urine_count,
            "排便回数": stool_count,
            **excretion_data,
            "家族共有メモ": family_memo,
            "気になる変化": changes,
            "登録日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "入力者": input_staff,
        }

        action = upsert_record(record)
        st.success(f"{action}しました。記録日＋利用者名をキーに、排泄詳細も保存されています。")


# =========================
# 過去データ管理
# 記録日＋利用者名を検索キーにした登録・検索・更新・削除
# =========================
elif menu == "過去データ管理":
    st.header("過去データ管理")
    st.caption("記録日＋利用者名を検索キーとして、データの検索・更新・削除を行います。")

    if st.session_state.role == "staff":
        st.info("お疲れ様です。入力内容の確認・修正が必要な場合はこちらから行えます。削除は慎重に行ってください。")

    df = load_data()
    df = normalize_key_columns(df)

    if df.empty:
        st.info("まだ登録データがありません。")
        st.stop()

    df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")

    st.subheader("検索キー")

    col1, col2, col3 = st.columns(3)

    with col1:
        key_date = st.date_input(
            "記録日",
            value=date.today(),
            key="past_key_date",
        )

    with col2:
        key_user = st.selectbox(
            "利用者名",
            all_users,
            key="past_key_user",
        )

    with col3:
        st.write("")
        st.write("")
        search_key_button = st.button("このキーで検索する", use_container_width=True)

    key_idx = find_record_index(df, key_date, key_user)

    if search_key_button:
        if key_idx is None:
            st.info("この記録日・利用者名のデータはありません。")
        else:
            st.success("該当データが見つかりました。下の更新フォームで編集できます。")

    st.divider()

    st.subheader("一覧検索")

    col4, col5, col6, col7 = st.columns(4)

    with col4:
        filter_user = st.selectbox(
            "利用者で絞り込み",
            ["全員"] + all_users,
            key="past_filter_user",
        )

    with col5:
        filter_year = st.number_input(
            "年",
            min_value=2024,
            max_value=2035,
            value=date.today().year,
            step=1,
            key="past_filter_year",
        )

    with col6:
        filter_month = st.number_input(
            "月",
            min_value=1,
            max_value=12,
            value=date.today().month,
            step=1,
            key="past_filter_month",
        )

    with col7:
        filter_day = st.selectbox(
            "日",
            ["全日"] + list(range(1, 32)),
            key="past_filter_day",
        )

    result = df.copy()
    result = result[
        (result["記録日"].dt.year == int(filter_year))
        & (result["記録日"].dt.month == int(filter_month))
    ]

    if filter_day != "全日":
        result = result[result["記録日"].dt.day == int(filter_day)]

    if filter_user != "全員":
        result = result[result["利用者名"] == filter_user]

    result = result.sort_values(["記録日", "利用者名"])

    st.write(f"該当件数：{len(result)}件")
    st.dataframe(result, use_container_width=True, hide_index=True)

    if not result.empty:
        excretion_cols = [
            "記録日",
            "利用者名",
            "排尿回数",
            "排便回数",
            "午前尿量",
            "午前尿性状",
            "午前便量",
            "午前便性状",
            "午後尿量",
            "午後尿性状",
            "午後便量",
            "午後便性状",
            "夕方尿量",
            "夕方尿性状",
            "夕方便量",
            "夕方便性状",
            "夜尿量",
            "夜尿性状",
            "夜便量",
            "夜便性状",
            "深夜尿量",
            "深夜尿性状",
            "深夜便量",
            "深夜便性状",
            "朝方尿量",
            "朝方尿性状",
            "朝方便量",
            "朝方便性状",
        ]
        visible_cols = [c for c in excretion_cols if c in result.columns]
        with st.expander("排泄詳細だけを確認する", expanded=False):
            st.dataframe(result[visible_cols], use_container_width=True, hide_index=True)

    st.divider()

    st.subheader("更新・削除")

    if key_idx is None:
        st.info("上の検索キーに該当するデータがないため、更新・削除フォームは表示されません。")
    else:
        selected_row = df.loc[key_idx]

        st.write("選択中の検索キー")
        st.code(f"{make_record_key(key_date, key_user)}")

        with st.form("key_update_form"):
            edit_date = st.date_input(
                "記録日",
                value=selected_row["記録日"].date() if pd.notna(selected_row["記録日"]) else key_date,
                key="edit_key_date",
            )

            edit_user = st.selectbox(
                "利用者名",
                all_users,
                index=all_users.index(selected_row["利用者名"]) if selected_row["利用者名"] in all_users else 0,
                key="edit_key_user",
            )

            st.markdown("#### バイタル")
            c1, c2, c3 = st.columns(3)
            with c1:
                edit_temp = st.number_input("体温", min_value=30.0, max_value=45.0, value=safe_float(selected_row["体温"], 36.5), step=0.1)
            with c2:
                edit_bp_high = st.number_input("血圧上", min_value=50, max_value=250, value=safe_int(selected_row["血圧上"], 120), step=1)
            with c3:
                edit_bp_low = st.number_input("血圧下", min_value=30, max_value=150, value=safe_int(selected_row["血圧下"], 75), step=1)

            c4, c5, c6 = st.columns(3)
            with c4:
                edit_pulse = st.number_input("脈拍", min_value=30, max_value=200, value=safe_int(selected_row["脈拍"], 70), step=1)
            with c5:
                edit_spo2 = st.number_input("SpO2", min_value=70, max_value=100, value=safe_int(selected_row["SpO2"], 96), step=1)
            with c6:
                edit_weight = st.number_input("体重", min_value=0.0, max_value=200.0, value=safe_float(selected_row["体重"], 50.0), step=0.1)

            st.markdown("#### 食事摂取率")
            m1, m2, m3 = st.columns(3)
            with m1:
                edit_breakfast = st.slider("朝食", 0, 100, safe_int(selected_row.get("朝食摂取率"), 80), step=10, key="edit_breakfast")
            with m2:
                edit_lunch = st.slider("昼食", 0, 100, safe_int(selected_row.get("昼食摂取率"), 80), step=10, key="edit_lunch")
            with m3:
                edit_dinner = st.slider("夕食", 0, 100, safe_int(selected_row.get("夕食摂取率"), 80), step=10, key="edit_dinner")

            st.markdown("#### 排泄詳細")
            edit_excretion_data = {}
            edit_urine_count = 0
            edit_stool_count = 0

            day_cols = st.columns(3)
            for col, slot_info in zip(day_cols, EXCRETION_SLOTS[:3]):
                slot, time_label = slot_info
                with col:
                    st.markdown(f"**{slot}**  \n{time_label}")
                    ua = st.selectbox(
                        f"{slot} 尿量",
                        URINE_AMOUNT_OPTIONS,
                        index=URINE_AMOUNT_OPTIONS.index(safe_text(selected_row.get(f"{slot}尿量", "なし"))) if safe_text(selected_row.get(f"{slot}尿量", "なし")) in URINE_AMOUNT_OPTIONS else 0,
                        key=f"edit_{slot}_ua",
                    )
                    ut = st.selectbox(
                        f"{slot} 尿性状",
                        URINE_TYPE_OPTIONS,
                        index=URINE_TYPE_OPTIONS.index(safe_text(selected_row.get(f"{slot}尿性状", "普通尿"))) if safe_text(selected_row.get(f"{slot}尿性状", "普通尿")) in URINE_TYPE_OPTIONS else 0,
                        key=f"edit_{slot}_ut",
                    )
                    sa = st.selectbox(
                        f"{slot} 便量",
                        STOOL_AMOUNT_OPTIONS,
                        index=STOOL_AMOUNT_OPTIONS.index(safe_text(selected_row.get(f"{slot}便量", "なし"))) if safe_text(selected_row.get(f"{slot}便量", "なし")) in STOOL_AMOUNT_OPTIONS else 0,
                        key=f"edit_{slot}_sa",
                    )
                    stt = st.selectbox(
                        f"{slot} 便性状",
                        STOOL_TYPE_OPTIONS,
                        index=STOOL_TYPE_OPTIONS.index(safe_text(selected_row.get(f"{slot}便性状", "普通便"))) if safe_text(selected_row.get(f"{slot}便性状", "普通便")) in STOOL_TYPE_OPTIONS else 0,
                        key=f"edit_{slot}_st",
                    )

                    edit_excretion_data[f"{slot}尿量"] = ua
                    edit_excretion_data[f"{slot}尿性状"] = "" if ua == "なし" else ut
                    edit_excretion_data[f"{slot}便量"] = sa
                    edit_excretion_data[f"{slot}便性状"] = "" if sa == "なし" else stt

                    if ua != "なし":
                        edit_urine_count += 1
                    if sa != "なし":
                        edit_stool_count += 1

            night_cols = st.columns(3)
            for col, slot_info in zip(night_cols, EXCRETION_SLOTS[3:]):
                slot, time_label = slot_info
                with col:
                    st.markdown(f"**{slot}**  \n{time_label}")
                    ua = st.selectbox(
                        f"{slot} 尿量",
                        URINE_AMOUNT_OPTIONS,
                        index=URINE_AMOUNT_OPTIONS.index(safe_text(selected_row.get(f"{slot}尿量", "なし"))) if safe_text(selected_row.get(f"{slot}尿量", "なし")) in URINE_AMOUNT_OPTIONS else 0,
                        key=f"edit_{slot}_ua",
                    )
                    ut = st.selectbox(
                        f"{slot} 尿性状",
                        URINE_TYPE_OPTIONS,
                        index=URINE_TYPE_OPTIONS.index(safe_text(selected_row.get(f"{slot}尿性状", "普通尿"))) if safe_text(selected_row.get(f"{slot}尿性状", "普通尿")) in URINE_TYPE_OPTIONS else 0,
                        key=f"edit_{slot}_ut",
                    )
                    sa = st.selectbox(
                        f"{slot} 便量",
                        STOOL_AMOUNT_OPTIONS,
                        index=STOOL_AMOUNT_OPTIONS.index(safe_text(selected_row.get(f"{slot}便量", "なし"))) if safe_text(selected_row.get(f"{slot}便量", "なし")) in STOOL_AMOUNT_OPTIONS else 0,
                        key=f"edit_{slot}_sa",
                    )
                    stt = st.selectbox(
                        f"{slot} 便性状",
                        STOOL_TYPE_OPTIONS,
                        index=STOOL_TYPE_OPTIONS.index(safe_text(selected_row.get(f"{slot}便性状", "普通便"))) if safe_text(selected_row.get(f"{slot}便性状", "普通便")) in STOOL_TYPE_OPTIONS else 0,
                        key=f"edit_{slot}_st",
                    )

                    edit_excretion_data[f"{slot}尿量"] = ua
                    edit_excretion_data[f"{slot}尿性状"] = "" if ua == "なし" else ut
                    edit_excretion_data[f"{slot}便量"] = sa
                    edit_excretion_data[f"{slot}便性状"] = "" if sa == "なし" else stt

                    if ua != "なし":
                        edit_urine_count += 1
                    if sa != "なし":
                        edit_stool_count += 1

            st.info(f"自動集計：排尿 {edit_urine_count} 回 ／ 排便 {edit_stool_count} 回")

            edit_family_memo = st.text_area("家族共有メモ", value=safe_text(selected_row["家族共有メモ"]))
            edit_changes = st.text_area("気になる変化", value=safe_text(selected_row["気になる変化"]))
            edit_staff = st.text_input("入力者", value=safe_text(selected_row["入力者"]))

            update_submit = st.form_submit_button("この内容で更新する")

        if update_submit:
            updated_record = {
                "記録日": edit_date,
                "利用者名": edit_user,
                "体温": edit_temp,
                "血圧上": edit_bp_high,
                "血圧下": edit_bp_low,
                "脈拍": edit_pulse,
                "SpO2": edit_spo2,
                "体重": edit_weight,
                "朝食摂取率": edit_breakfast,
                "昼食摂取率": edit_lunch,
                "夕食摂取率": edit_dinner,
                "排尿回数": edit_urine_count,
                "排便回数": edit_stool_count,
                **edit_excretion_data,
                "家族共有メモ": edit_family_memo,
                "気になる変化": edit_changes,
                "登録日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "入力者": edit_staff,
            }

            action = upsert_record(updated_record)
            st.success(f"{action}しました。")
            st.rerun()

        st.subheader("削除")
        st.warning("削除すると元に戻せません。")

        delete_check = st.checkbox("この検索キーのデータを削除することを確認しました")

        if st.button("このデータを削除する"):
            if not delete_check:
                st.error("削除する場合は確認チェックを入れてください。")
            else:
                original_df = load_data()
                delete_idx = find_record_index(original_df, key_date, key_user)

                if delete_idx is None:
                    st.error("削除対象が見つかりません。")
                else:
                    original_df = original_df.drop(index=delete_idx).reset_index(drop=True)
                    save_data(original_df)
                    st.success("削除しました。")
                    st.rerun()


# =========================
# 入力データ確認
# =========================
elif menu == "入力データ確認":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("入力データ確認")
    st.caption("管理者用：利用者・年・月・日で入力データを絞り込みできます。")

    df = load_data()

    filter_options = ["全員"] + all_users

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        filter_user = st.selectbox("利用者で絞り込み", filter_options)

    with col2:
        year = st.number_input(
            "年",
            min_value=2024,
            max_value=2035,
            value=date.today().year,
            step=1,
        )

    with col3:
        month = st.number_input(
            "月",
            min_value=1,
            max_value=12,
            value=date.today().month,
            step=1,
        )

    with col4:
        day_filter = st.selectbox(
            "日",
            ["全日"] + list(range(1, 32)),
        )

    view = df.copy()

    if not view.empty:
        view["記録日"] = pd.to_datetime(view["記録日"], errors="coerce")

        view = view[
            (view["記録日"].dt.year == int(year))
            & (view["記録日"].dt.month == int(month))
        ]

        if day_filter != "全日":
            view = view[view["記録日"].dt.day == int(day_filter)]

        if filter_user != "全員":
            view = view[view["利用者名"] == filter_user]

        view = view.sort_values(["記録日", "利用者名"], ascending=[True, True])

    st.subheader("検索結果")

    if view.empty:
        st.warning("該当するデータがありません。")
    else:
        st.write(f"該当件数：{len(view)}件")
        st.dataframe(view, use_container_width=True, hide_index=True)

    with open(DATA_FILE, "rb") as f:
        st.download_button(
            "入力データExcelをダウンロード",
            data=f,
            file_name="健康チェック入力データ.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# =========================
# 家族向けレポート作成
# =========================
elif menu == "家族向けレポート作成":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("家族向けレポート作成")
    df = load_data()

    if not all_users:
        st.warning("利用者が登録されていません。")
        st.stop()

    col1, col2, col3 = st.columns(3)

    with col1:
        report_user = st.selectbox("利用者名", all_users)

    with col2:
        report_year = st.number_input("対象年", min_value=2024, max_value=2035, value=date.today().year, step=1)

    with col3:
        report_month = st.number_input("対象月", min_value=1, max_value=12, value=date.today().month, step=1)

    if st.button("家族向けレポートを作成"):
        report_path, target, summary_text = create_family_report(df, report_user, report_year, report_month)

        if target.empty:
            st.warning("指定した利用者・年月のデータがありません。空のレポートを作成しました。")
        else:
            st.success("家族向けレポートを作成しました。")

        st.subheader("AI家族レポート文章プレビュー")
        st.info(summary_text)

        with open(report_path, "rb") as f:
            st.download_button(
                "作成したレポートをダウンロード",
                data=f,
                file_name=report_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )



# =========================
# ひだまりレポートPDF
# =========================
elif menu == "ひだまりレポートPDF":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("ひだまりレポートPDF")
    st.caption("ご家族にお渡ししやすい、イラスト風の表紙・グラフ・月間まとめ付きPDFを作成します。")

    df = load_data()

    if not all_users:
        st.warning("利用者が登録されていません。")
        st.stop()

    col1, col2, col3 = st.columns(3)

    with col1:
        pdf_user = st.selectbox("利用者名", all_users, key="pdf_user")

    with col2:
        pdf_year = st.number_input(
            "対象年",
            min_value=2024,
            max_value=2035,
            value=date.today().year,
            step=1,
            key="pdf_year",
        )

    with col3:
        pdf_month = st.number_input(
            "対象月",
            min_value=1,
            max_value=12,
            value=date.today().month,
            step=1,
            key="pdf_month",
        )

    st.info("PDFには、今月のまとめ・健康状態の目安・バイタル推移グラフ・家族共有メモ・気になる変化が入ります。")

    if st.button("ひだまりレポートPDFを作成する"):
        try:
            pdf_path, summary_text, target = create_hidamari_report_pdf(
                df,
                pdf_user,
                pdf_year,
                pdf_month,
            )

            if target.empty:
                st.warning("指定した利用者・年月のデータがありません。空のレポートを作成しました。")
            else:
                st.success("ひだまりレポートPDFを作成しました。")

            st.subheader("レポート文章プレビュー")
            st.info(summary_text)

            with open(pdf_path, "rb") as f:
                st.download_button(
                    "PDFをダウンロード",
                    data=f,
                    file_name=pdf_path.name,
                    mime="application/pdf",
                )

        except ModuleNotFoundError as e:
            st.error("PDF作成に必要なライブラリが不足しています。requirements.txt に reportlab と matplotlib を追加してください。")
            st.code("reportlab\nmatplotlib")
        except Exception as e:
            st.error("PDF作成中にエラーが発生しました。")
            st.exception(e)



# =========================
# 管理者支援
# =========================
elif menu == "管理者支援":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("管理者支援")
    st.caption("AI家族レポート、バイタル推移グラフ、ChatGPT連携用プロンプト、申し送り支援をまとめています。")

    df = load_data()

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        [
            "AI家族レポート",
            "バイタル推移グラフ",
            "ChatGPT連携",
            "申し送り支援",
            "アセスメント分析",
        ]
    )

    with tab1:
        st.subheader("AI家族レポート自動文章")
        if not all_users:
            st.warning("利用者が登録されていません。")
        else:
            col1, col2, col3 = st.columns(3)
            with col1:
                ai_user = st.selectbox("利用者", all_users, key="ai_user")
            with col2:
                ai_year = st.number_input("対象年", min_value=2024, max_value=2035, value=date.today().year, step=1, key="ai_year")
            with col3:
                ai_month = st.number_input("対象月", min_value=1, max_value=12, value=date.today().month, step=1, key="ai_month")

            target = get_month_data(df, ai_user, ai_year, ai_month)
            summary = create_family_summary_text(target, ai_user, ai_year, ai_month)
            st.text_area("家族向け文章", value=summary, height=300)

            st.subheader("基軸となるアセスメント情報")
            st.text_area(
                "AIが参照する背景情報",
                value=build_assessment_context_text(ai_user),
                height=220,
            )

    with tab2:
        st.subheader("バイタル推移グラフ")
        if df.empty:
            st.info("データがありません。")
        else:
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                graph_user = st.selectbox("利用者", all_users, key="graph_user")
            with col2:
                graph_item = st.selectbox("項目", ["体温", "血圧上", "血圧下", "脈拍", "SpO2", "体重", "朝食摂取率", "昼食摂取率", "夕食摂取率", "排尿回数", "排便回数"], key="graph_item")
            with col3:
                graph_year = st.number_input("年", min_value=2024, max_value=2035, value=date.today().year, step=1, key="graph_year")
            with col4:
                graph_month = st.number_input("月", min_value=1, max_value=12, value=date.today().month, step=1, key="graph_month")

            target = get_month_data(df, graph_user, graph_year, graph_month)
            if target.empty:
                st.warning("対象データがありません。")
            else:
                chart_df = target[["記録日", graph_item]].copy()
                chart_df[graph_item] = pd.to_numeric(chart_df[graph_item], errors="coerce")
                chart_df = chart_df.dropna()
                chart_df = chart_df.set_index("記録日")
                st.line_chart(chart_df)

                st.dataframe(target, use_container_width=True, hide_index=True)

    with tab3:
        st.subheader("ChatGPT連携用プロンプト")
        st.caption("このアプリ内では外部API接続は行わず、ChatGPTに貼り付ける文章を作成します。個人情報の扱いには注意してください。")

        if df.empty:
            st.info("データがありません。")
        else:
            prompt_user = st.selectbox("利用者", all_users, key="prompt_user")
            prompt_year = st.number_input("対象年", min_value=2024, max_value=2035, value=date.today().year, step=1, key="prompt_year")
            prompt_month = st.number_input("対象月", min_value=1, max_value=12, value=date.today().month, step=1, key="prompt_month")

            target = get_month_data(df, prompt_user, prompt_year, prompt_month)

            prompt_text = f"""あなたは介護施設の家族向けレポートを整える文章整理係です。
以下の健康チェック記録をもとに、ご家族へ渡す月間レポート文を作成してください。

【重要ルール】
・医療判断、診断、治療効果の断定はしない。
・「問題ありません」「改善しました」「安心です」と断定しない。
・『記録上、大きな変化は見られていません』『様子を見守ります』のように、記録に基づく表現にする。
・ご家族が読みやすい、やわらかく丁寧な文章にする。

【対象】
利用者名：{prompt_user}
対象月：{prompt_year}年{prompt_month}月

【基軸となるアセスメント情報】
{build_assessment_context_text(prompt_user)}

【記録データ】
{target.to_string(index=False)}
"""
            st.text_area("ChatGPTに貼り付けるプロンプト", value=prompt_text, height=420)

    with tab4:
        st.subheader("申し送り支援")
        handover_date = st.date_input("申し送り対象日", value=date.today())
        handover = create_handover_text(df, handover_date)
        st.text_area("申し送りまとめ", value=handover, height=360)


    with tab5:
        st.subheader("アセスメント分析")
        st.caption("利用者マスタのアセスメント情報と月間記録を組み合わせ、管理者が確認すべき視点を整理します。")

        if not all_users:
            st.warning("利用者が登録されていません。")
        else:
            col1, col2, col3 = st.columns(3)
            with col1:
                assess_user = st.selectbox("利用者", all_users, key="assess_user")
            with col2:
                assess_year = st.number_input(
                    "対象年",
                    min_value=2024,
                    max_value=2035,
                    value=date.today().year,
                    step=1,
                    key="assess_year",
                )
            with col3:
                assess_month = st.number_input(
                    "対象月",
                    min_value=1,
                    max_value=12,
                    value=date.today().month,
                    step=1,
                    key="assess_month",
                )

            assess_target = get_month_data(df, assess_user, assess_year, assess_month)
            analysis_text = build_admin_assessment_analysis(assess_user, assess_target)

            st.text_area("管理者確認メモ", value=analysis_text, height=420)

            prompt_text = f"""あなたは介護施設の管理者支援を行う文章整理係です。
以下のアセスメント情報と月間記録をもとに、職員間で共有しやすい確認ポイントを整理してください。

【重要ルール】
・医療判断、診断、治療効果の断定はしない。
・本人や職員を責める表現にしない。
・事実、背景、確認すべき点、支援の方向性を分ける。
・ADL、IADL、認知機能、食事、排泄、気になる変化を関連づけて整理する。

【利用者】
{assess_user}

【アセスメント情報】
{build_assessment_context_text(assess_user)}

【月間記録】
{assess_target.to_string(index=False)}
"""
            st.text_area("ChatGPT連携用プロンプト", value=prompt_text, height=460)



# =========================
# 排泄詳細管理
# =========================
elif menu == "排泄詳細管理":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("排泄詳細管理")
    st.caption("日中帯・夜間帯の排泄詳細を集計し、管理者が状況を把握するための画面です。")

    df = load_data()

    if df.empty:
        st.info("まだ登録データがありません。")
        st.stop()

    col1, col2, col3 = st.columns(3)

    with col1:
        ex_user = st.selectbox("利用者", ["全員"] + all_users, key="ex_user")

    with col2:
        start_date = st.date_input(
            "開始日",
            value=date.today(),
            key="ex_start_date",
        )

    with col3:
        end_date = st.date_input(
            "終了日",
            value=date.today(),
            key="ex_end_date",
        )

    target_users = active_users if ex_user == "全員" else [ex_user]

    summary_df, detail_df, alert_df = build_excretion_admin_summary(
        df,
        target_users,
        start_date,
        end_date,
    )

    st.subheader("排泄サマリー")

    if summary_df.empty:
        st.warning("該当する排泄データがありません。")
    else:
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

    st.subheader("注意して確認したい排泄記録")

    if alert_df.empty:
        st.success("指定期間内に、濃縮尿・下痢便・水様便などの注意記録はありません。")
    else:
        st.warning("確認したい排泄記録があります。")
        st.dataframe(alert_df, use_container_width=True, hide_index=True)

    st.subheader("時系列の排泄詳細")

    if detail_df.empty:
        st.info("排泄詳細データはありません。")
    else:
        st.dataframe(detail_df, use_container_width=True, hide_index=True)

        csv = detail_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "排泄詳細CSVをダウンロード",
            data=csv,
            file_name="排泄詳細データ.csv",
            mime="text/csv",
        )

    st.divider()
    st.subheader("管理者向け確認メモ")

    if summary_df.empty:
        st.info("確認メモを作成できるデータがありません。")
    else:
        memo_lines = [
            "排泄詳細データをもとにした管理者確認メモです。",
            "医療判断ではなく、職員間の共有と見守り方針の整理に使用してください。",
            "",
        ]

        for _, row in summary_df.iterrows():
            memo_lines.append(
                f"■ {row['利用者名']}：排尿{row['排尿回数合計']}回、排便{row['排便回数合計']}回、"
                f"排便なし日数{row['排便なし日数']}日。{row['注意メモ']}"
            )

        if not alert_df.empty:
            memo_lines.append("")
            memo_lines.append("【確認したい記録】")
            for _, row in alert_df.head(10).iterrows():
                date_text = row["記録日"].strftime("%m/%d") if pd.notna(row["記録日"]) else ""
                memo_lines.append(f"・{date_text} {row['利用者名']}：{row['確認内容']}")

        st.text_area(
            "確認メモ",
            value="\\n".join(memo_lines),
            height=320,
        )



# =========================
# 利用者マスタ管理
# =========================
elif menu == "利用者マスタ管理":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("利用者マスタ管理")
    st.caption("利用者の追加・非表示・アセスメント情報の登録管理ができます。")

    df_users = load_users(include_hidden=True)

    st.subheader("現在の利用者一覧")
    st.dataframe(df_users, use_container_width=True, hide_index=True)

    st.divider()

    st.subheader("利用者を追加")
    with st.form("add_user_form", clear_on_submit=True):
        new_user = st.text_input("追加する利用者名", placeholder="例：田中様")
        add_submit = st.form_submit_button("追加する")

    if add_submit:
        ok, msg = add_user(new_user)

        if ok:
            st.success(msg)
            st.rerun()
        else:
            st.error(msg)

    st.divider()

    st.subheader("アセスメント情報の登録・更新")
    st.caption("家族レポートやAI分析の背景情報として活用できます。")

    if df_users.empty:
        st.info("利用者が登録されていません。")
    else:
        selected_user = st.selectbox(
            "アセスメントを編集する利用者",
            df_users["利用者名"].tolist(),
            key="assessment_edit_user",
        )

        selected_df = df_users[df_users["利用者名"] == selected_user]

        if not selected_df.empty:
            selected = selected_df.iloc[0]

            with st.form("assessment_form"):
                basic_info = st.text_area(
                    "基本情報（氏名・住所など）",
                    value=safe_text(selected.get("基本情報", "")),
                    height=80,
                    placeholder="例：年齢、家族構成、生活上の基本情報など",
                )

                main_complaint = st.text_area(
                    "主訴（本人・家族の希望や困りごと）",
                    value=safe_text(selected.get("主訴", "")),
                    height=100,
                    placeholder="例：本人・家族が困っていること、希望していること",
                )

                life_status = st.text_area(
                    "生活状況（1日の流れ）",
                    value=safe_text(selected.get("生活状況", "")),
                    height=120,
                    placeholder="例：起床、食事、排泄、入浴、就寝などの流れ",
                )

                adl = st.text_area(
                    "ADL（日常生活動作）",
                    value=safe_text(selected.get("ADL", "")),
                    height=100,
                    placeholder="例：移動、食事、更衣、排泄、入浴など。条件付きで記載",
                )

                iadl = st.text_area(
                    "IADL（生活関連動作）",
                    value=safe_text(selected.get("IADL", "")),
                    height=100,
                    placeholder="例：買い物、服薬管理、金銭管理、掃除など。「していない」と「できない」を区別",
                )

                cognitive = st.text_area(
                    "認知機能（判断・記憶）",
                    value=safe_text(selected.get("認知機能", "")),
                    height=100,
                    placeholder="例：生活場面で見られる判断・記憶の様子",
                )

                health = st.text_area(
                    "健康状態（疾患・服薬）",
                    value=safe_text(selected.get("健康状態", "")),
                    height=100,
                    placeholder="例：疾患、服薬、生活への影響など",
                )

                issue = st.text_area(
                    "課題（支援が必要な問題点）",
                    value=safe_text(selected.get("課題", "")),
                    height=100,
                    placeholder="例：背景や原因もあわせて整理",
                )

                support = st.text_area(
                    "支援内容（具体的な対応）",
                    value=safe_text(selected.get("支援内容", "")),
                    height=100,
                    placeholder="例：実行可能な支援内容",
                )

                assessment_submit = st.form_submit_button("アセスメント情報を保存する")

            if assessment_submit:
                df_save = load_users(include_hidden=True)

                mask = df_save["利用者名"] == selected_user

                if mask.any():
                    df_save.loc[mask, "基本情報"] = basic_info
                    df_save.loc[mask, "主訴"] = main_complaint
                    df_save.loc[mask, "生活状況"] = life_status
                    df_save.loc[mask, "ADL"] = adl
                    df_save.loc[mask, "IADL"] = iadl
                    df_save.loc[mask, "認知機能"] = cognitive
                    df_save.loc[mask, "健康状態"] = health
                    df_save.loc[mask, "課題"] = issue
                    df_save.loc[mask, "支援内容"] = support

                    save_users(df_save)

                    st.success("アセスメント情報を保存しました。")
                    st.rerun()
                else:
                    st.error("対象の利用者が見つかりません。")

    st.divider()

    st.subheader("利用者を入力候補から外す")
    visible_users = load_users(include_hidden=False)["利用者名"].tolist()

    if visible_users:
        target_user = st.selectbox("対象利用者", visible_users)
        st.warning("この操作は、入力画面の候補から外すだけです。過去データとアセスメント情報は削除されません。")

        if st.button("入力候補から外す"):
            ok, msg = hide_user(target_user)

            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    else:
        st.info("現在、表示中の利用者はいません。")

    st.subheader("非表示の利用者を戻す")
    df_users = load_users(include_hidden=True)
    hidden_df = df_users[df_users["表示"] == "非表示"]

    if not hidden_df.empty:
        restore_user = st.selectbox("表示に戻す利用者", hidden_df["利用者名"].tolist())

        if st.button("表示に戻す"):
            ok, msg = add_user(restore_user)

            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    else:
        st.info("非表示の利用者はいません。")

    with open(USER_FILE, "rb") as f:
        st.download_button(
            "利用者マスタExcelをダウンロード",
            data=f,
            file_name="利用者マスタ.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
